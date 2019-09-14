# Fetch data from 'website_name', clean the data and store it in excel sheet(xls) 
import json
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

try:
	# load the sheet if it has already been created
	wb = load_workbook('1_sheet_name.xlsx')
	companies_sheet = wb.active
except:
	# Create object of Workbook
	wb = Workbook()
	# 'create_sheet' is used to create sheet. 
	companies_sheet = wb.create_sheet('1_sheet_name')

# Total number of companies
total_city_companies = companies_sheet.max_row - 2
print(total_city_companies)

# 'i' if for each row (outer loop) - each company
i = 2
# 'j' is for each column (inner loop). Each company has 'n' fields where n > 1
j = 2

# Base URL for the website.
BASE_URL = "base_url"

# Set this to the city name for which you want the data. Like Mumbai etc...
CITY_NAME = "city_name"

for scrip_code in range(500000,600000):
    
    # 'company_name_string' object will contain the data received.
    company_name_string = requests.get("company_name_string_api").text
    # 'company_name_string_json' will make the json parsable. 
    company_name_string_json = json.loads(company_name_string)
    
    # 'company_data_string' object will contain the data received.
    company_data_string = requests.get("company_data_string_api").text
    # 'company_data_string_json' will make the json parsable. 
    company_data_string_json = json.loads(company_data_string)
   
    if company_name_string_json["Cmpname"]["FullN"]:
        if company_data_string_json["Table1"]:
            if company_data_string_json["Table1"][0]["City"] == CITY_NAME:
                
                total_city_companies = total_city_companies + 1
                
                companies_sheet.cell(i,j, company_name_string_json["Cmpname"]["FullN"])
                j = j + 1
                companies_sheet.cell(i,j, company_name_string_json["Cmpname"]["ShortN"])
                j = j + 1
                companies_sheet.cell(i,j, scrip_code)
                j = j + 1
                companies_sheet.cell(i,j, company_name_string_json["Cmpname"]["Category"])
                j = j + 1
                companies_sheet.cell(i,j, BASE_URL + company_name_string_json["Cmpname"]["SEOUrlEQ"])
                j = j + 1
                
                if company_data_string_json["Table"]:
                    company_secretary_data = company_data_string_json["Table"][len(company_data_string_json["Table"]) - 1]
                    companies_sheet.cell(i,j, company_secretary_data["sPrefix"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_secretary_data["sFirstname"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_secretary_data["sMiddlename"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_secretary_data["sLastname"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_secretary_data["sDesignation"])
                    j = j + 1
                else: 
                    j = j + 5
                
                
                if company_data_string_json["Table1"]:
                    company_data = company_data_string_json["Table1"][0]
                    companies_sheet.cell(i,j, company_data["Address"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["City"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["nPIN"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["State"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["Tele"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["Fax"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["sEmail"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_data["sURL"])
                    j = j + 1
                else:
                    j + j + 8
                
                
                if company_data_string_json["Table2"]:
                    company_reg_data = company_data_string_json["Table2"][0]
                    companies_sheet.cell(i,j, company_reg_data["fld_scripcode"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_reg_data["RegName"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_reg_data["Address"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_reg_data["Phone"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_reg_data["Fax"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_reg_data["fld_Emailid"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_reg_data["WebSite"])
                    j = j + 1
                else:
                    j = j + 7
                
                if company_data_string_json["Table3"]:
                    company_industry_data = company_data_string_json["Table3"][0]
                    companies_sheet.cell(i,j, company_industry_data["ISIN_NUMBER"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_industry_data["Industry"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_industry_data["Impact_Cost"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_industry_data["BCRD"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_industry_data["MARKET_LOT"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_industry_data["lISTING_DATE"])
                    j = j + 1
                    companies_sheet.cell(i,j, company_industry_data["fld_cin"])
                else: 
                    j = j + 6
                
                # Store the "city_name" company in the excel sheet
                wb.save('1_sheet_name.xlsx') 
            
                i = i + 1
                j = 2
                print("Scrip Code:" + str(scrip_code) + "  Total " + CITY_NAME + " Companies:" + str(total_city_companies))
            else: 
                print("Scrip Code:" + str(scrip_code) + " false")
        else:
            print("Scrip Code:" + str(scrip_code) + " false")
    else:
        print("Scrip Code:" + str(scrip_code) + " false")

i = i + 1
companies_sheet.write(i,j, "Total Companies In " + CITY_NAME)
companies_sheet.write(i,j + 1, total_city_companies)
# Store the "city_name" company count in the excel sheet
wb.save('1_sheet_name.xls')
